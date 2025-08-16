from openpyxl import Workbook, load_workbook

Tech_data = load_workbook(r"C:\Users\ebunope.ajayi\OneDrive - IBEDC\Desktop\IBEDC 2025\TECH DATA FOLDER FROM PRECIOUS SYSTEM\TECH DATA\TECHNICAL DATA AS AT\2025\JULY 2025\TECHNICAL DATA 21-07-2025.xlsx")


realignment_data = load_workbook()

Tech_data_active = Tech_data.active
realignment_data_active = realignment_data.active

